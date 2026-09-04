#!/usr/bin/env python3
"""
Excel报表导出脚本 - 从SQLite读取购买数据，用openpyxl填充模板，保存到输出文件
用法: python3 export_excel.py <template_path> <output_path> <db_path> <purchase_id> <user_id> [comp_screenshot] [shop_screenshot] [map_screenshot]
"""
import sys
import json
import sqlite3
import os
from openpyxl import load_workbook, styles
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO

def main():
    if len(sys.argv) < 6:
        print(json.dumps({"error": "参数不足: template_path output_path db_path purchase_id user_id"}))
        sys.exit(1)

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    db_path = sys.argv[3]
    purchase_id = int(sys.argv[4])
    user_id = int(sys.argv[5])

    import math

    # Haversine 公式计算距离（米）
    def haversine(lat1, lng1, lat2, lng2):
        R = 6371000
        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        delta_phi = math.radians(lat2 - lat1)
        delta_lambda = math.radians(lng2 - lng1)
        a = math.sin(delta_phi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(delta_lambda/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

    try:
        # 从数据库读取购买记录（保持连接打开供后续查询竞品使用）
        db = sqlite3.connect(db_path)
        db.row_factory = sqlite3.Row
        row = db.execute(
            "SELECT * FROM purchases WHERE id = ? AND user_id = ?",
            (purchase_id, user_id)
        ).fetchone()

        if not row:
            print(json.dumps({"error": "记录不存在"}))
            sys.exit(1)

        row = dict(row)
        store_name = row.get('store_name', '门店')
        radius = row.get('radius', 500)
        city_month = row.get('city_month', '')
        center_lat = row.get('center_lat', 0)
        center_lng = row.get('center_lng', 0)

        # 解析 result_data
        result_data = row.get('result_data')
        if isinstance(result_data, str):
            result_data = json.loads(result_data)
        api_result = result_data.get('apiResult', {}) if result_data else {}

        # 半径转km
        if isinstance(radius, str):
            radii_list = json.loads(radius)
        elif isinstance(radius, (list, tuple)):
            radii_list = radius
        else:
            radii_list = [radius]
        actual_radius_meters = radii_list[0] if radii_list else 3000
        radius_km = actual_radius_meters / 1000
        radius_str = f"{actual_radius_meters}米" if actual_radius_meters < 1000 else f"{radius_km}km"

        # 打开模板
        wb = load_workbook(template_path)

        # ===== 1. 封面 =====
        if '封面' in wb.sheetnames:
            ws = wb['封面']
            ws['E21'] = store_name

        # ===== 2. 商圈数据 =====
        if '商圈数据' in wb.sheetnames:
            ws = wb['商圈数据']
            ws['C1'] = radius_str
            ws['E1'] = city_month

            # ===== 构建按服务码分区的字段映射 =====
            # v1.13.97 修正：旧实现用全局 field_map（键=字段名小写），1009/1010/1011/1012/1013
            # 五服务区的同名字段 pop{群}_p{档} 互相覆盖（后遍历的 1013 覆盖消费力/教育/行业/人生四区）。
            # 现按「成品服务码」分区 field_map[code][field_lower]，模板行按其 A 列 code 归属取对应分区。
            field_map = {}  # code -> {field_lower: value}

            def _put(code, key, val):
                try:
                    val = int(round(float(val)))
                except (TypeError, ValueError):
                    return
                field_map.setdefault(code, {})[str(key).lower()] = val

            # 1001 - 对象
            if '1001' in api_result and isinstance(api_result['1001'], dict):
                for k, v in api_result['1001'].items():
                    _put('1001', k, v)

            # 1002 - 标签（按人群类型分开：0=到访 1=居住 2=工作）
            if '1002' in api_result and isinstance(api_result['1002'], list):
                tag_names = ['网上购物', '时政要闻', '商务办公', '金融理财', '手机游戏', '旅游出行', '外卖送餐', '餐饮美食', '求职招聘']
                for item in api_result['1002']:
                    name = item.get('tag_name', '')
                    val = item.get('tag_value', 0)
                    pop = item.get('popu_type', 0)
                    # 1. 精确匹配优先（如"求职招聘"，避免被子串"招聘"覆盖）
                    if name in tag_names:
                        idx = tag_names.index(name)
                        _put('1002', f'webtag{pop}_{idx + 1}', val)
                        continue
                    # 2. 包含匹配（如"时政要闻"等），已写过的槽位不覆盖
                    for i, tn in enumerate(tag_names, 1):
                        if tn in name or name in tn:
                            key = f'webtag{pop}_{i}'
                            if key not in field_map.setdefault('1002', {}):
                                _put('1002', key, val)
                            break

            # 1005 - 小时段
            if '1005' in api_result and isinstance(api_result['1005'], list):
                for item in api_result['1005']:
                    dt = item.get('day_type', 0)
                    hp = item.get('hour_period')
                    if hp is not None:
                        _put('1005', f'hour{dt}_{hp}_visit', item.get('hour_visit', 0))
                        _put('1005', f'hour{dt}_{hp}_all', item.get('hour_all', 0))

            # 1006 - 日均
            if '1006' in api_result and isinstance(api_result['1006'], list) and len(api_result['1006']) > 0:
                days = api_result['1006']
                n = len(days)
                keys = ['day_visit', 'day_all', 'stay1', 'stay2', 'stay3', 'stay4', 'stay5']
                totals = {k: 0 for k in keys}
                for d in days:
                    for k in keys:
                        totals[k] += d.get(k, 0) or 0
                _put('1006', 'day_avg_visit', round(totals['day_visit'] / n))
                _put('1006', 'day_avg_total', round(totals['day_all'] / n))
                _put('1006', 'stay_30', round(totals['stay1'] / n))
                _put('1006', 'stay_60', round(totals['stay2'] / n))
                _put('1006', 'stay_120', round(totals['stay3'] / n))
                _put('1006', 'stay_240', round(totals['stay4'] / n))
                _put('1006', 'stay_480', round(totals['stay5'] / n))

            # 1007 每月到达次数（真实返回元素无 popu_type、直接含 reach1~5；旧 popu_type==0 判断会漏配）
            if '1007' in api_result and isinstance(api_result['1007'], list):
                reach_item = next(
                    (it for it in api_result['1007'] if isinstance(it, dict) and any(str(k).startswith('reach') for k in it)),
                    None
                )
                if reach_item:
                    for i in range(1, 6):
                        _put('1007', f'reach{i}', reach_item.get(f'reach{i}', 0) or 0)

            # 1009 消费水平（富裕指数）：元素 spendpower='1'..'8' + spendpower_value（模板 pop{群}_p{档}）
            if '1009' in api_result and isinstance(api_result['1009'], list):
                for item in api_result['1009']:
                    pt = item.get('popu_type', 0)
                    sp = item.get('spendpower')
                    if sp is not None:
                        _put('1009', f'pop{pt}_p{sp}', item.get('spendpower_value', 0))

            # 1010 教育 / 1011 行业 / 1012 人生阶段 / 1013 综合消费能力：各自独立分区
            for svc in ['1010', '1011', '1012', '1013']:
                if svc in api_result and isinstance(api_result[svc], list):
                    for item in api_result[svc]:
                        pt = item.get('popu_type', 0)
                        for k, v in item.items():
                            if k == 'popu_type':
                                continue
                            if k.startswith('p') and isinstance(v, (int, float)):
                                _put(svc, f'pop{pt}_{k}', v or 0)

            # 1015 资产预测数组（收入/有车/有房 fname 分流 → 成品报表区 1014/1015/1016）
            if '1015' in api_result and isinstance(api_result['1015'], list):
                fname_map = {'收入预测': ('1014', 'income'), '有车预测': ('1015', 'car'), '有房预测': ('1016', 'house')}
                for item in api_result['1015']:
                    fm = fname_map.get(item.get('fname', ''))
                    if not fm:
                        continue
                    dst_code, prefix = fm
                    pt = item.get('popu_type', 0)
                    for k, v in item.items():
                        if k == 'popu_type':
                            continue
                        if k.startswith('p') and isinstance(v, (int, float)):
                            _put(dst_code, f'{prefix}_pop{pt}_{k}', v or 0)

            # 填入E列（按模板行 A 列 code 继承归属取对应分区）
            cur_code = None
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
                code_val = row[0].value
                if code_val not in (None, ''):
                    cur_code = str(code_val).strip()
                field_name = str(row[2].value or '').strip().lower()
                if not field_name or cur_code is None:
                    continue
                seg = field_map.get(cur_code)
                if seg and field_name in seg:
                    ws.cell(row=row[0].row, column=5).value = seg[field_name]

        # ===== 3. 竞品列表（中心3km范围内） =====
        if '竞品列表' in wb.sheetnames:
            ws = wb['竞品列表']

            # 查询当前用户的竞品门店
            competitors = db.execute(
                "SELECT * FROM competitors WHERE user_id = ?",
                (user_id,)
            ).fetchall()

            # 计算距离并筛选3km内
            nearby = []
            for comp in competitors:
                comp = dict(comp)
                dist = haversine(center_lat, center_lng, comp['latitude'], comp['longitude'])
                if dist <= 3000:  # 3km内
                    comp['distance'] = round(dist, 1)
                    nearby.append(comp)

            # 按距离排序
            nearby.sort(key=lambda c: c['distance'])

            # ---- 品牌汇总：按品牌分组计数，写入 C3:J4（8个槽位，品牌奇数列/数量偶数列） ----
            from collections import Counter
            brand_counter = Counter()
            for comp in nearby:
                b = comp.get('brand') or '未知'
                brand_counter[b] += 1
            top_brands = brand_counter.most_common(8)  # [(品牌, 数量)]，按数量降序

            brand_cells = ['C3', 'E3', 'G3', 'I3', 'C4', 'E4', 'G4', 'I4']
            count_cells = ['D3', 'F3', 'H3', 'J3', 'D4', 'F4', 'H4', 'J4']
            for i, (bname, bcount) in enumerate(top_brands):
                ws[brand_cells[i]] = bname
                ws[count_cells[i]] = bcount
            # 剩余槽位清空，避免旧数据残留
            for i in range(len(top_brands), 8):
                ws[brand_cells[i]] = None
                ws[count_cells[i]] = None

            # ---- 竞品明细：从第7行开始输出（第3-4行为品牌汇总区，第5-6行预留） ----
            # 列映射：A=店名, B=地址, C=分类, D=价格, E=星级, F=评论数, G=口味, H=环境, I=服务, J=距离
            row_idx = 7  # 从第7行开始输出
            for comp in nearby:
                ws.cell(row=row_idx, column=1, value=comp.get('name', ''))
                ws.cell(row=row_idx, column=2, value=comp.get('address', ''))
                ws.cell(row=row_idx, column=3, value=comp.get('store_category', ''))
                ws.cell(row=row_idx, column=4, value=comp.get('price', 0))
                ws.cell(row=row_idx, column=5, value=comp.get('rating', 0))
                ws.cell(row=row_idx, column=6, value=int(comp.get('reviews', 0)))
                ws.cell(row=row_idx, column=7, value=comp.get('taste_score', 0))
                ws.cell(row=row_idx, column=8, value=comp.get('environment_score', 0))
                ws.cell(row=row_idx, column=9, value=comp.get('service_score', 0))
                ws.cell(row=row_idx, column=10, value=round(comp['distance']))
                row_idx += 1

        # ===== 3.5 购物中心列表（中心3km范围内） =====
        if '购物中心列表' in wb.sheetnames:
            ws_shopping = wb['购物中心列表']
        else:
            ws_shopping = wb.create_sheet('购物中心列表')
            ws_shopping['A3'] = '名称'
            ws_shopping['B3'] = '地址'
            ws_shopping['C3'] = '星级'
            ws_shopping['D3'] = '评论数'
            ws_shopping['E3'] = '榜单'
            ws_shopping['F3'] = '距离(米)'

        # 查询购物中心（不过滤user_id，与前端API /shopping-centers-for-map 行为一致）
        shopping_centers = db.execute(
            "SELECT * FROM shopping_centers"
        ).fetchall()

        # 计算距离并筛选3km内
        nearby_centers = []
        for sc in shopping_centers:
            sc = dict(sc)
            dist = haversine(center_lat, center_lng, sc['latitude'], sc['longitude'])
            if dist <= 3000:
                sc['distance'] = round(dist, 1)
                nearby_centers.append(sc)

        nearby_centers.sort(key=lambda c: c['distance'])

        row_idx = 4
        for sc in nearby_centers:
            ws_shopping.cell(row=row_idx, column=1, value=sc.get('name', ''))
            ws_shopping.cell(row=row_idx, column=2, value=sc.get('address', ''))
            ws_shopping.cell(row=row_idx, column=3, value=sc.get('stars', ''))
            ws_shopping.cell(row=row_idx, column=4, value=sc.get('comments', 0))
            ws_shopping.cell(row=row_idx, column=5, value=sc.get('rank_info', ''))
            ws_shopping.cell(row=row_idx, column=6, value=round(sc['distance']))
            row_idx += 1

        # 关闭数据库
        db.close()

        # ===== 4. 地图 + 竞品地图 + 购物中心地图（如有截图） =====
        comp_screenshot = sys.argv[6] if len(sys.argv) >= 7 else None
        shop_screenshot = sys.argv[7] if len(sys.argv) >= 8 else None
        map_screenshot = sys.argv[8] if len(sys.argv) >= 9 else None

        # 通用函数：在指定sheet的A3插入截图
        def embed_screenshot(sheet_name, file_path, img_w=1200, img_h=786):
            if not file_path or not os.path.exists(file_path):
                return
            try:
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
                img = XLImage(file_path)
                img.width = img_w
                img.height = img_h
                ws.add_image(img, 'A3')
            except Exception as e:
                print(f'[EXPORT_WARN] {sheet_name}截图嵌入失败: {e}')

        embed_screenshot('竞品地图', comp_screenshot, 1200, 786)
        embed_screenshot('购物中心地图', shop_screenshot, 1200, 786)
        embed_screenshot('地图', map_screenshot, 1200, 786)

        # 保存
        wb.save(output_path)

        # 保存完成后清理截图文件
        for fp in [comp_screenshot, shop_screenshot, map_screenshot]:
            if fp and os.path.exists(fp):
                try: os.remove(fp)
                except: pass

        # 输出文件名信息（给Node.js使用）
        file_name = f"{store_name}_{radius_str}_{city_month}.xlsx"
        print(json.dumps({"success": True, "filename": file_name}))

    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()
